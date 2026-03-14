from typing import Type

import csv

import inspect

from os import path

from typing import List

import pythoncom
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from concurrent.futures import ThreadPoolExecutor
import win32com.client


class Excel:
    allowed_file_type = ('.xlsx', '.xlsm', '.xlsb', '.xlt', '.xltm', '.xls', '.xlt')

    @staticmethod
    def checkExcelFileIsOpened(full_file_path: str) -> None:
        """Verifica se um arquivo do Excel já está aberto."""
        try:
            pythoncom.CoInitialize()
            try:
                excel = win32com.client.GetActiveObject("Excel.Application")
            except Exception:
                # Nenhuma instância ativa do Excel foi encontrada
                return

            try:
                workbooks = excel.Workbooks
            except Exception:
                raise ExcelWorkbooksAccessError(
                    classe=__class__.__name__,
                    metodo=inspect.currentframe().f_code.co_name
                )

            for wb in workbooks:
                if wb.FullName.lower() == full_file_path.lower():
                    raise WorkBookAlreadyOpened(
                        workbook_path=full_file_path,
                        classe=__class__.__name__,
                        metodo=inspect.currentframe().f_code.co_name
                    )
        except WorkBookAlreadyOpened as wbo:
            raise wbo
        finally:
            pythoncom.CoUninitialize()

    @staticmethod
    def createWorkbook(file_path: str) -> Workbook:
        """Cria uma pplanilha Excel com o nome informado e retorna o objeto workbook.
        :param file_path: str do caminho completo do arquivo a ser criado.
        :return: Workbook
        :raises: Erro não mapeado ao criar arquivo Excel
        """
        # Excel.checkFullFilePath(file_name)
        if not file_path.lower().endswith(Excel.allowed_file_type):
            raise NotExcelFile(file_path=file_path,
                               classe=__class__.__name__,
                               metodo=inspect.currentframe().f_code.co_name)
        try:
            workbook = Workbook()
            workbook.save(filename=file_path)
            return workbook
        except Exception as e:
            raise Exception(f"Erro não mapeado ao criar arquivo Excel.\n"
                            f"Classe: {__class__.__name__}\n"
                            f"Método: {inspect.currentframe().f_code.co_name}\n"
                            f"Exceção:\n{str(e)}\n"
                            # f"Traceback:\n {str(traceback.format_exc())}"
                            )

    @staticmethod
    def readWorkBook(file_path: str, read_only=True, data_only=True):
        """Lê um arquivo Excel e armazena em self._workbook.
        :param file_path: str = Caminho completo do arquivo Excel.
        :param read_only: bool = True para abrir o arquivo em modo somente leitura.
        :param data_only: bool = True para ler apenas os valores das células.
        :return: bool = True se o arquivo foi lido com sucesso.
        """
        try:
            return load_workbook(filename=file_path, read_only=read_only, data_only=data_only)
        except Exception as e:
            raise Exception(f"Erro não mapeado ao ler arquivo Excel.\n"
                            f"Classe: {__class__.__name__}\n"
                            f"Método: {inspect.currentframe().f_code.co_name}\n"
                            f"Exceção:\n{str(e)}\n"
                            # f"Traceback:\n {str(traceback.format_exc())}"
                            )

    @staticmethod
    def createWorksheet(workbook: Workbook, worksheet_name: str, layout: Type['WorkSheetLayout'] = None,
                        header_row: int = 1) -> Worksheet:

        """Cria uma aba, com o nome informado, em um objeto workbook passado como parâmetro.
        :param workbook: Workbook
        :param worksheet_name: str
        :param layout: LayoutExcel
        :param header_row: int = Número da linha onde estão os nomes das colunas.
        :return:  Worksheet
        :raises: WorkSheetAlreadyExists
        :raises: Erro não mapeado ao criar arquivo Excel
        """
        try:
            if worksheet_name in [name for name in Excel.getWorkSheetsList(workbook)]:
                raise WorkSheetAlreadyExists(
                    worksheet_name=worksheet_name,
                    classe=Excel.__class__.__name__,
                    metodo=inspect.currentframe().f_code.co_name,
                    workbook_path=workbook.path
                )
            workbook.create_sheet(worksheet_name)
            worksheet = workbook[worksheet_name]
            if layout is not None:
                for key, fields in layout.copyEstrutura().items():
                    worksheet[fields[layout.col_destino] + str(header_row)] = fields[layout.col_nome]
            return worksheet

        except WorkSheetAlreadyExists as wsa:
            raise wsa

        except Exception as e:
            raise Exception(f"Erro não mapeado ao criar aba no arquivo Excel.\n"
                            f"Classe: {__class__.__name__}\n"
                            f"Método: {inspect.currentframe().f_code.co_name}\n"
                            f"Exceção:\n{str(e)}\n"
                            # f"Traceback:\n {str(traceback.format_exc())}"
                            )

    @staticmethod
    def getWorkSheetsList(workbook: Workbook) -> List[str]:
        """Retorna uma lista com os nomes das planilhas do arquivo Excel.
        :param workbook: Workbook
        :return: List[str]
        :raises: WorkbookNotFound
        """
        if not workbook:
            raise WorkbookNotFound(workbook_path='',
                                   classe=__class__.__name__,
                                   metodo=inspect.currentframe().f_code.co_name)
        return workbook.sheetnames

    @staticmethod
    def checkWorkSheetExist(workbook: Workbook, worksheet_name: str) -> bool:
        """Verifica se a planilha existe no arquivo Excel.
        :param workbook: Workbook
        :param worksheet_name: str
        :return: bool
        """
        if worksheet_name not in [name for name in Excel.getWorkSheetsList(workbook)]:
            return False
        return True

    @staticmethod
    def returnFloatValue(value) -> float:
        try:
            return float(str(value))
        except:
            return 0.0

    @staticmethod
    def checkFullFilePath(full_file_path: str):
        f"""Verifica se o caminho informado é um arquivo Excel. Tipos aceitos: {Excel.allowed_file_type}.
        :param full_file_path: str
        :raises: NotAFile, NotExcelFile
        """

        if not path.isfile(full_file_path.replace('\\\\', '\\')):
            raise NotAFile(file_path=full_file_path,
                           classe=__class__.__name__,
                           metodo=inspect.currentframe().f_code.co_name)
        if not full_file_path.lower().endswith(Excel.allowed_file_type):
            raise NotExcelFile(file_path=full_file_path,
                               classe=__class__.__name__,
                               metodo=inspect.currentframe().f_code.co_name)

    @staticmethod
    def writeWorkSheet(workbook: Workbook, worksheet_name: str, dict_layout_data: dict, num_threads: int = 4):
        """Cria uma planilha no arquivo Excel utilizando threads para paralelizar a escrita.
        :param workbook: Workbook
        :param worksheet_name: str
        :param dict_layout_data: dict = {id_unico (row_number): {'chave_do_campo': {'col_destino': 'A', 'valor': '}}}
        :param num_threads: int = Número de threads a serem usadas.
        :raises: WorkSheetAlreadyExists
        """
        try:
            first_row = 1
            if not Excel.checkWorkSheetExist(workbook, worksheet_name):
                workbook.create_sheet(worksheet_name)
                worksheet = workbook[worksheet_name]
                # Preenche o cabeçalho
                for row, dict_data in dict_layout_data.items():
                    for campo, props in dict_data.items():
                        posicao = props['col_destino'] + '1'
                        worksheet[posicao] = str(props['col_nome'])
                    first_row = 2
                    break
            else:
                worksheet = workbook[worksheet_name]
                first_row = worksheet.max_row + 1

            # Função para processar e escrever uma linha
            def process_row(cont_row, key, dict_data):
                for campo, props in dict_data.items():
                    posicao = props['col_destino'] + str(first_row + cont_row)
                    worksheet[posicao] = props['valor'] if props['valor'] and str(
                        props['valor']).upper() != 'NONE' else ''
                    worksheet[posicao].number_format = props['formato'] if 'formato' in props else 'General'

            # Usa ThreadPoolExecutor para escrever as linhas em paralelo
            with ThreadPoolExecutor(max_workers=num_threads) as executor:
                futures = [executor.submit(process_row, cont_row, key, dict_data)
                           for cont_row, (key, dict_data) in enumerate(dict_layout_data.items())]

            # Coleta os resultados
            for future in futures:
                future.result()  # Se houver exceções, elas serão capturadas aqui

        except Exception as e:
            raise Exception(f"Erro não mapeado ao criar aba no arquivo Excel.\n"
                            f"Classe: {__class__.__name__}\n"
                            f"Método: {inspect.currentframe().f_code.co_name}\n"
                            f"Exceção:\n{str(e)}\n")

    @staticmethod
    def writeWorkSheetBK(workbook: Workbook, worksheet_name: str, dict_layout_data: dict):
        """Cria uma planilha no arquivo Excel.
        :param workbook: Workbook
        :param worksheet_name: str
        :param dict_layout_data: dict = {id_unico (row_number): {'chave_do_campo': {'col_destino': 'A', 'valor': '}}}
        :param workbook_path: str = Caminho completo do arquivo Excel.
        :raises: WorkSheetAlreadyExists
        """
        try:
            first_row = 1
            if not Excel.checkWorkSheetExist(workbook, worksheet_name):
                workbook.create_sheet(worksheet_name)
                worksheet = workbook[worksheet_name]
                for row, dict_data in dict_layout_data.items():
                    for campo, props in dict_data.items():
                        posicao = props['col_destino'] + '1'
                        worksheet[posicao] = str(props['col_nome'])
                    first_row = 2
                    break
            else:
                worksheet = workbook[worksheet_name]
                first_row = worksheet.max_row + 1

            for cont_row, (key, dict_data) in enumerate(dict_layout_data.items()):
                for campo, props in dict_data.items():
                    posicao = props['col_destino'] + str(first_row + cont_row)
                    worksheet[posicao] = props['valor'] if props['valor'] and str(
                        props['valor']).upper() != 'NONE' else ''
                    worksheet[posicao].number_format = props['formato'] if 'formato' in props else 'General'

            # # Pega o intervalo dinâmico preenchido da worksheet
            # tabela_intervalo = worksheet.dimensions
            #
            # # Cria uma tabela com base no intervalo dinâmico
            # tabela = Table(displayName=f"tabela_{worksheet_name.replace(' ', '_')}", ref=tabela_intervalo)
            #
            # # Define um estilo para a tabela (opcional)
            # estilo = TableStyleInfo(
            #     name="TableStyleMedium9",  # Estilo predefinido
            #     showFirstColumn=False,
            #     showLastColumn=False,
            #     showRowStripes=True,  # Linhas listradas
            #     showColumnStripes=False
            # )
            #
            # # Aplica o estilo à tabela
            # tabela.tableStyleInfo = estilo
            #
            # # Adiciona a tabela à worksheet
            # worksheet.add_table(tabela)

        except Exception as e:
            raise Exception(f"Erro não mapeado ao criar aba no arquivo Excel.\n"
                            f"Classe: {__class__.__name__}\n"
                            f"Método: {inspect.currentframe().f_code.co_name}\n"
                            f"Exceção:\n{str(e)}\n"
                            # f"Traceback:\n {str(traceback.format_exc())}"
                            )

    @staticmethod
    def writeWorkSheetRow(worksheet: Worksheet, row_number: int, layout: Type['WorkSheetLayout'] = None,
                          create_header=False):
        """Cria uma planilha no arquivo Excel.
        :param worksheet: Worksheet = Planilha do arquivo Excel.
        :param row_number: int = Número da linha a ser preenchida.
        :param layout: LayoutExcel, exemplo: {'chave_do_campo': {'col_destino': 'A', 'valor': '}}
        :param create_header: bool = True para criar o cabeçalho da planilha.
        :raises: WorkSheetAlreadyExists
        """
        try:
            if create_header:
                for key, dict_data in layout.estrutura.items():
                    posicao = dict_data[layout.col_destino] + '1'
                    worksheet[posicao] = str(dict_data[layout.col_nome])

            for key, dict_data in layout.estrutura.items():
                posicao = dict_data[layout.col_destino] + str(row_number)
                worksheet[posicao] = dict_data[layout.valor] if dict_data[layout.valor] and str(
                    dict_data[layout.valor]).upper() != 'NONE' else ''
                worksheet[posicao].number_format = dict_data[layout.formato] \
                    if layout.formato in dict_data else 'General'

        except Exception as e:
            raise Exception(f"Erro não mapeado ao criar aba no arquivo Excel.\n"
                            f"Layout: {layout.__class__.__name__}"
                            f"Classe: {__class__.__name__}\n"
                            f"Método: {inspect.currentframe().f_code.co_name}\n"
                            f"Exceção:\n{str(e)}\n"
                            # f"Traceback:\n {str(traceback.format_exc())}"
                            )

    @staticmethod
    def writeWorkSheetRowbk2(worksheet: Worksheet, row_number: int, layout: Type['WorkSheetLayout'] = None,
                          create_header=False):
        """Cria uma planilha no arquivo Excel.
        :param worksheet: Worksheet = Planilha do arquivo Excel.
        :param row_number: int = Número da linha a ser preenchida.
        :param layout: LayoutExcel, exemplo: {'chave_do_campo': {'col_destino': 'A', 'valor': '}}
        :param create_header: bool = True para criar o cabeçalho da planilha.
        :raises: WorkSheetAlreadyExists
        """
        try:
            dict_layout_data = layout.copyEstrutura()
            if create_header:
                for key, dict_data in dict_layout_data.items():
                    posicao = dict_data[layout.col_destino] + '1'
                    worksheet[posicao] = str(dict_data[layout.col_nome])

            for key, dict_data in dict_layout_data.items():
                posicao = dict_data[layout.col_destino] + str(row_number)
                worksheet[posicao] = dict_data[layout.valor] if dict_data[layout.valor] and str(
                    dict_data[layout.valor]).upper() != 'NONE' else ''
                worksheet[posicao].number_format = dict_data[layout.formato] \
                    if layout.formato in dict_data else 'General'

        except Exception as e:
            raise Exception(f"Erro não mapeado ao criar aba no arquivo Excel.\n"
                            f"Layout: {layout.__class__.__name__}"
                            f"Classe: {__class__.__name__}\n"
                            f"Método: {inspect.currentframe().f_code.co_name}\n"
                            f"Exceção:\n{str(e)}\n"
                            # f"Traceback:\n {str(traceback.format_exc())}"
                            )

    @staticmethod
    def writeWorkSheetRowBK(worksheet: Worksheet, row_number: int, dict_layout_data: dict,
                          create_header=False):
        """Cria uma planilha no arquivo Excel.
        :param worksheet: Worksheet = Planilha do arquivo Excel.
        :param row_number: int = Número da linha a ser preenchida.
        :param dict_layout_data: dict = {id_unico (row_number): {'chave_do_campo': {'col_destino': 'A', 'valor': '}}}
        :param create_header: bool = True para criar o cabeçalho da planilha.
        :raises: WorkSheetAlreadyExists
        """
        try:
            if create_header:
                for key, dict_data in dict_layout_data.items():
                    posicao = dict_data['col_destino'] + '1'
                    worksheet[posicao] = str(dict_data['col_nome'])

            for key, dict_data in dict_layout_data.items():
                posicao = dict_data['col_destino'] + str(row_number)
                worksheet[posicao] = dict_data['valor'] if dict_data['valor'] and str(
                    dict_data['valor']).upper() != 'NONE' else ''
                worksheet[posicao].number_format = dict_data['formato'] if 'formato' in dict_data else 'General'

        except Exception as e:
            raise Exception(f"Erro não mapeado ao criar aba no arquivo Excel.\n"
                            f"Classe: {__class__.__name__}\n"
                            f"Método: {inspect.currentframe().f_code.co_name}\n"
                            f"Exceção:\n{str(e)}\n"
                            # f"Traceback:\n {str(traceback.format_exc())}"
                            )

    @staticmethod
    def writeWorkSheet2(workbook: Workbook, worksheet_name: str, dict_layout_data: dict):
        """Cria uma planilha no arquivo Excel.
        :param workbook: Workbook
        :param worksheet_name: str
        :param dict_layout_data: dict = {id_unico (row_number): {'chave_do_campo': {'col_destino': 'A', 'valor': '}}}
        :param workbook_path: str = Caminho completo do arquivo Excel.
        :raises: WorkSheetAlreadyExists
        """
        try:
            first_row = 1
            if not Excel.checkWorkSheetExist(workbook, worksheet_name):
                workbook.create_sheet(worksheet_name)
                worksheet = workbook[worksheet_name]
                for row, dict_data in dict_layout_data.items():
                    for campo, props in dict_data.items():
                        posicao = props['col_destino'] + '1'
                        worksheet[posicao] = str(props['col_nome'])
                    first_row = 2
                    break
            else:
                worksheet = workbook[worksheet_name]
                first_row = worksheet.max_row + 1

            for cont_row, (key, dict_data) in enumerate(dict_layout_data.items()):
                for campo, props in dict_data.items():
                    posicao = props['col_destino'] + str(first_row + cont_row)
                    worksheet[posicao] = props['valor'] if props['valor'] and str(
                        props['valor']).upper() != 'NONE' else ''
                    worksheet[posicao].number_format = props['formato'] if 'formato' in props else 'General'

            # # Pega o intervalo dinâmico preenchido da worksheet
            # tabela_intervalo = worksheet.dimensions
            #
            # # Cria uma tabela com base no intervalo dinâmico
            # tabela = Table(displayName=f"tabela_{worksheet_name.replace(' ', '_')}", ref=tabela_intervalo)
            #
            # # Define um estilo para a tabela (opcional)
            # estilo = TableStyleInfo(
            #     name="TableStyleMedium9",  # Estilo predefinido
            #     showFirstColumn=False,
            #     showLastColumn=False,
            #     showRowStripes=True,  # Linhas listradas
            #     showColumnStripes=False
            # )
            #
            # # Aplica o estilo à tabela
            # tabela.tableStyleInfo = estilo
            #
            # # Adiciona a tabela à worksheet
            # worksheet.add_table(tabela)

        except Exception as e:
            raise Exception(f"Erro não mapeado ao criar aba no arquivo Excel.\n"
                            f"Classe: {__class__.__name__}\n"
                            f"Método: {inspect.currentframe().f_code.co_name}\n"
                            f"Exceção:\n{str(e)}\n"
                            # f"Traceback:\n {str(traceback.format_exc())}"
                            )



class Conversor:

    @staticmethod
    def ExcelToCSV(workbook: Workbook, worksheet_name: str, csv_filename: str, delimitador: str = ','):
        ws = workbook[worksheet_name]
        with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
            csvwriter = csv.writer(csvfile, delimiter=delimitador)
            for row in ws.iter_rows(values_only=True):
                csvwriter.writerow(row)

    @staticmethod
    def CSVToExcel(csv_filename: str, excel_filename: str, sheet_name: str = 'Sheet1', delimitador: str = ','):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        with open(csv_filename, 'r', encoding='utf-8') as csvfile:
            csvreader = csv.reader(csvfile, delimiter=delimitador)
            for row in csvreader:
                ws.append(row)

        wb.save(excel_filename)

class NotAFile(Exception):
    def __init__(self, file_path: str, classe: str = '', metodo: str = ''):
        self.message = f"O caminho informado não é um arquivo: '{file_path}'\n" \
                       f"{'Classe: ' + classe if classe else ''}\n" \
                       f"{'Método: ' + metodo if metodo else ''}"
        super().__init__(self.message)

class NotExcelFile(Exception):
    def __init__(self, file_path: str = '', classe: str = '', metodo: str = ''):
        self.message = f"O arquivo passado não é um arquivo excel.\n" \
                       f"Tipos de arquivos permitidos: {Excel.allowed_file_type}.\n" \
                       f"Caminho informado: '{file_path}'" \
                       f"{'Classe: ' + classe if classe else ''}\n" \
                       f"{'Método: ' + metodo if metodo else ''}"

        super().__init__(self.message)

class WorkbookNotFound(Exception):
    def __init__(self,
                 workbook_path: str,
                 classe: str,
                 metodo: str
                 ):
        self.message = (f"Objeto Workbook não encontrado.\n"
                        f"Verifique se o arquivo foi lido corretamente: '{workbook_path}'"
                        f"{'Classe: ' + classe if classe else ''}\n"
                        f"{'Método: ' + metodo if metodo else ''}"
                        )
        super().__init__(self.message)

class WorkSheetAlreadyExists(Exception):
    def __init__(self,
                 workbook_path: str,
                 worksheet_name: str,
                 classe: str = '',
                 metodo: str = ''
                 ):
        self.message = (f"WorSheet já existe: '{worksheet_name}'"
                        f"No arquivo: '{workbook_path}'.\n"
                        f"{'Classe: ' + classe if classe else ''}\n"
                        f"{'Método: ' + metodo if metodo else ''}"
                        )
        super().__init__(self.message)

class WorkBookAlreadyOpened(Exception):
    def __init__(self,
                 workbook_path: str,
                 classe: str = '',
                 metodo: str = ''
                 ):
        self.message = (f"O arquivo está aberto: '{workbook_path}'.\n"
                        f"Feche-o e reinicie a automação.\n"
                        f"{'Classe: ' + classe if classe else ''}\n"
                        f"{'Método: ' + metodo if metodo else ''}"
                        )
        super().__init__(self.message)

class ExcelWorkbooksAccessError(Exception):
    def __init__(self,
                 classe: str = '',
                 metodo: str = ''
                 ):
        self.message = (
            "Não foi possível acessar a lista de workbooks da instância ativa do Excel.\n"
            "Verifique se o Excel está funcionando corretamente ou se há permissões de acesso suficientes.\n"
            f"{'Classe: ' + classe if classe else ''}"
            f"{' | Método: ' + metodo if metodo else ''}"
        )
        super().__init__(self.message)


if __name__ == '__main__':
    from Projeto.Scripts.Auxiliar.Auxiliar import Auxiliar
    workbook = Excel.readWorkBook(f'{Auxiliar.getProjectRootDir()}\\Bases de Testes\\Requições de Compra20-05-2025 14-53-02.xlsx')
    # Excel.checkExcelFileIsOpened(r'C:\Users\rafael.m.cavalcante\PycharmProjects\BatimentosCPFL\ResultadoBatimento.xlsx')
    Conversor.ExcelToCSV(
        workbook=workbook,
        worksheet_name='Requisições',
        csv_filename=f'{Auxiliar.getProjectRootDir()}\\Bases de Testes\\Testes.csv',
        delimitador=','
    )