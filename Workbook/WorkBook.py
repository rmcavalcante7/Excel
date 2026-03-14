import inspect

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

from Excel import Excel
from Worksheet.WorksheetLayout import WorkSheetLayout


class WorkBook:

    def __init__(self, file_path: str, read_only: bool = True, data_only: bool = True):
        self.excel = Excel
        self.file_path = file_path
        self.workbook = self.excel.readWorkBook(file_path=file_path, read_only=read_only, data_only=data_only)
        # self._workbook: [Workbook | None] = None
        self._active_worksheet = None
        self._worksheets: dict = {str: Worksheet}  # {'nome_planilha': Worksheet}
        self.work_sheet_list = []

    @property
    def file_path(self):
        return self.__file_path

    @file_path.setter
    def file_path(self, file_path: str):
        self.excel.checkFullFilePath(file_path)
        self.__file_path = file_path

    @property
    def worksheets(self):
        return self._worksheets

    def readWorkSheet(self, worksheet_name: str) -> Worksheet:
        """Lê uma planilha do arquivo Excel e armazena em self._worksheet.
        :param worksheet_name: str
        :return: Worksheet
        :raises: WorkbookNotFound, WorkSheetNotFound
        :raises: Erro não mapeado ao ler worksheet.
        """
        try:
            if not self.workbook:
                raise WorkbookNotFound(workbook_path=self.file_path,
                                       classe=self.__class__.__name__,
                                       metodo=inspect.currentframe().f_code.co_name)

            if worksheet_name is None or str(worksheet_name).upper() == 'NONE' or str(worksheet_name).strip() == '':
                raise WorkSheetNotFound(workbook_path=self.file_path,
                                        worksheet_name=str(worksheet_name),
                                        classe=self.__class__.__name__,
                                        metodo=inspect.currentframe().f_code.co_name
                                        )

            if not self.checkWorkSheetExist(worksheet_name):
                raise WorkSheetNotFound(workbook_path=self.file_path,
                                        worksheet_name=worksheet_name,
                                        classe=self.__class__.__name__,
                                        metodo=inspect.currentframe().f_code.co_name
                                        )

            self._worksheets[worksheet_name] = self.workbook[worksheet_name]
            self._active_worksheet = self._worksheets[worksheet_name]
            return self._active_worksheet

        except WorkbookNotFound as wbnf:
            raise wbnf

        except WorkSheetNotFound as wsnf:
            raise wsnf

        except Exception as e:
            raise Exception(f"Erro não mapeado ao ler worksheet.\n"
                            f"Classe: {self.__class__.__name__}\n"
                            f"Método: {inspect.currentframe().f_code.co_name}\n"
                            f"Exceção:\n{str(e)}\n"
                            )

    def getWorkSheetsList(self):
        return Excel.getWorkSheetsList(self.workbook)

    def checkWorkSheetExist(self, worksheet_name):
        return Excel.checkWorkSheetExist(self.workbook, worksheet_name)

    def createWorkSheet(self, worksheet_name: str, dict_layout_data: dict, dict_layout: WorkSheetLayout):
        """Cria uma planilha no arquivo Excel.
        :param worksheet_name: str
        :param dict_layout_data: Dicionário onde a chave é o número da linha e o valor é um dicionário com o
                                WorkSheetLayout preenchido
        :param dict_layout: WorkSheetLayout
        :raises: WorkSheetAlreadyExists
        """
        try:
            self.workbook.create_sheet(worksheet_name)
            self.worksheets[worksheet_name] = self.workbook[worksheet_name]
            worksheet = self.worksheets[worksheet_name]
            for cont_row, (row_id, dict_data) in enumerate(dict_layout_data.items()):
                for campo, props in dict_data.items():
                    if cont_row == 0:
                        worksheet[props['col_final'] + str(cont_row + 1)] = props['nome_final']

                    worksheet[props['col_final'] + str(cont_row + 2)] = props['value']
                    if props['pattern_fill']:
                        start_color = props['pattern_fill']['start_color'] if props['pattern_fill']['start_color'] else None
                        end_color = props['pattern_fill']['end_color'] if props['pattern_fill']['end_color'] else None
                        fill_type = props['pattern_fill']['fill_type'] if props['pattern_fill']['fill_type'] else None
                        fill = PatternFill(start_color=start_color, end_color=end_color, fill_type=fill_type)
                        worksheet[props['col_final'] + str(cont_row + 2)].fill = fill

            self.workbook.save(self.file_path)
        except Exception as e:
            raise Exception(f"Erro não mapeado ao criar aba no arquivo Excel.\n"
                            f"Classe: {self.__class__.__name__}\n"
                            f"Método: {inspect.currentframe().f_code.co_name}\n"
                            f"Exceção:\n{str(e)}\n"
                            # f"Traceback:\n {str(traceback.format_exc())}"
                            )

    def createWorkSheetOLD(self, worksheet_name: str, dict_layout_data: dict, dict_layout: WorkSheetLayout):
        """Cria uma planilha no arquivo Excel.
        :param worksheet_name: str
        :param dict_layout_data: dict = {id_unico (row_number): {'chave_do_campo': {'col_destino': 'A', 'valor': '}}}
        :raises: WorkSheetAlreadyExists
        """
        try:
            if not self.checkWorkSheetExist(worksheet_name):
                self.workbook.create_sheet(worksheet_name)
                self.worksheets[worksheet_name] = self.workbook[worksheet_name]

            worksheet = self.worksheets[worksheet_name]
            first_row = worksheet.min_row
            for row_, dict_data in dict_layout_data.items():
                for campo, props in dict_data.items():
                    if campo in dict_layout.lista_exececao:
                        continue
                    worksheet[props['col_final'] + str(first_row + row_)] = props['value']

            self.workbook.save(self.file_path)
        except Exception as e:
            raise Exception(f"Erro não mapeado ao criar aba no arquivo Excel.\n"
                            f"Classe: {self.__class__.__name__}\n"
                            f"Método: {inspect.currentframe().f_code.co_name}\n"
                            f"Exceção:\n{str(e)}\n"
                            # f"Traceback:\n {str(traceback.format_exc())}"
                            )

    def vLookup(self, worksheet_name, value_to_find, column_to_search,
                column_to_return, case_sensitive: bool = False):
        self.readWorkSheet(worksheet_name)
        worksheet = self._worksheets[worksheet_name]

        value_to_find = str(value_to_find)
        row_max = worksheet.max_row
        for row_ in range(1, row_max + 1):
            value = worksheet[column_to_search + str(row_)].value
            value = str(value)
            if value is not None:
                if not case_sensitive:
                    if value.strip().upper() == value_to_find.upper():
                        value_to_return = worksheet[column_to_return + str(row_)]
                        if value_to_return is None:
                            return None
                        return value_to_return.value
                else:
                    if value.strip() == value_to_find:
                        value_to_return = worksheet[column_to_return + str(row_)]
                        if value_to_return is None:
                            return None
                        return value_to_return.value
        return None

    def vLookupAll(self, worksheet_name, value_to_find, column_to_search, columns_to_return: list,
                   case_sensitive: bool = False):
        self.readWorkSheet(worksheet_name)
        worksheet = self._worksheets[worksheet_name]

        return_list = []
        value_to_find = str(value_to_find)
        row_max = worksheet.max_row
        for row_ in range(1, row_max + 1):
            value = worksheet[column_to_search + str(row_)].value
            value = str(value)
            if value is not None:
                if not case_sensitive:
                    if value.strip().upper() == value_to_find.upper():
                        for column_to_return in columns_to_return:
                            value_to_return = worksheet[column_to_return + str(row_)]
                            if value_to_return is None:
                                return_list.append(None)
                            else:
                                return_list.append(value_to_return.value)
                        return return_list
                else:
                    if value.strip() == value_to_find:
                        for column_to_return in columns_to_return:
                            value_to_return = worksheet[column_to_return + str(row_)]
                            if value_to_return is None:
                                return_list.append(None)
                            else:
                                return_list.append(value_to_return.value)
                        return return_list
        return []

    def hLookup(self, worksheet_name, value_to_find, row_to_search, row_to_return, case_sensitive: bool = False):
        self.readWorkSheet(worksheet_name)
        worksheet = self._worksheets[worksheet_name]

        value_to_find = str(value_to_find)
        column_max = worksheet.max_column
        for column in range(1, column_max + 1):
            column_letter = get_column_letter(column)
            value = worksheet[column_letter + str(row_to_search)].value
            value = str(value)
            if value is not None:
                if not case_sensitive:
                    if value.strip().upper() == value_to_find.upper():
                        value_to_return = worksheet[column_letter + str(row_to_return)]
                        if value_to_return is None:
                            return None
                        return value_to_return.value
                else:
                    if value.strip() == value_to_find:
                        value_to_return = worksheet[column_letter + str(row_to_return)]
                        if value_to_return is None:
                            return None
                        return value_to_return.value
        return None

    def hLookupAll(self, worksheet_name, value_to_find, row_to_search, rows_to_return, case_sensitive: bool = False):
        self.readWorkSheet(worksheet_name)
        worksheet = self._worksheets[worksheet_name]

        return_list = []
        value_to_find = str(value_to_find)
        column_max = worksheet.max_column
        for column in range(1, column_max + 1):
            column_letter = get_column_letter(column)
            value = worksheet[column_letter + str(row_to_search)].value
            value = str(value)
            if value is not None:
                if not case_sensitive:
                    if value.strip().upper() == value_to_find.upper():
                        for row_to_return in rows_to_return:
                            value_to_return = worksheet[column_letter + str(row_to_return)]
                            if value_to_return is None:
                                return_list.append(None)
                            else:
                                return_list.append(value_to_return.value)
                        return return_list
                else:
                    if value.strip() == value_to_find:
                        for row_to_return in rows_to_return:
                            value_to_return = worksheet[column_letter + str(row_to_return)]
                            if value_to_return is None:
                                return_list.append(None)
                            else:
                                return_list.append(value_to_return.value)
                        return return_list
        return []

    def writeToWorkSheetFromColumns(self, worksheet_name, dict_data: dict, save_file=False):
        """As keys de Dict_data são as colunas da planilha e os VALUES serão os valores a serem preenchidos na coulna."""
        # print('teste')
        if self.workbook is None:
            return False

        self.readWorkSheet(worksheet_name)
        worksheet = self._worksheets[worksheet_name]

        try:
            for column_number, key in enumerate(dict_data.keys()):

                column_letter = get_column_letter(column_number + 1)
                for row_number, value in enumerate(dict_data[key]):
                    worksheet[column_letter + str(row_number + 2)] = value
            if save_file:
                self.workbook.save(self.file_path)
            return True
        except Exception as e:
            print(str(e))
            return False

    def writeToWorkSheetFromRows(self, dict_data: dict, save_file=False):
        """As keys de Dict_data são as linhas da planilha e os VALUES serão os valores a serem preenchidos na linha."""
        # print('teste')
        if self.workbook is None:
            return False

        if self._active_worksheet is None:
            return False

        try:
            if len(dict_data) > 0:
                if isinstance(dict_data[list(dict_data.keys())[0]], dict):
                    row_number = 0
                    for key in dict_data.keys():
                        inner_dict = dict_data[key]
                        for lista_dados in inner_dict.values():
                            for column_number, value in enumerate(lista_dados):
                                column_letter = get_column_letter(column_number + 1)
                                self._active_worksheet[column_letter + str(row_number + 2)] = value
                            row_number += 1

                else:
                    # ainda precisa ser implementado
                    for row_number, key in enumerate(dict_data.keys()):
                        for column_number, value in enumerate(dict_data[key]):
                            column_letter = get_column_letter(column_number + 1)
                            self._active_worksheet[column_letter + str(row_number + 2)] = value

            if save_file:
                self.workbook.save(self.file_path)
            return True
        except Exception as e:
            print(str(e))
            return False

    def closeWorkBook(self):
        try:
            if self.workbook is not None:
                self.workbook.close()
                self.workbook = None
                return True
            else:
                return False
        except Exception as e:
            print(e)


## classe para um path informado que não seja um arquivo
class NotAFile(Exception):
    def __init__(self, file_path: str, classe: str = '', metodo: str = ''):
        self.message = f"O caminho informado não é um arquivo: '{file_path}'\n" \
                       f"{'Classe: ' + classe if classe else ''}\n" \
                       f"{'Método: ' + metodo if metodo else ''}"
        super().__init__(self.message)


class NotExcelFile(Exception):
    def __init__(self, file_path: str = '', classe: str = '', metodo: str = ''):
        self.message = f"O arquivo passado não é um arquivo excel.\n" \
                       f"Tipos de arquivos permitidos: {Excel.allowed_file_type}.\n" \
                       f"Caminho informado: '{file_path}'\n" \
                       f"{'Classe: ' + classe if classe else ''}\n" \
                       f"{'Método: ' + metodo if metodo else ''}"

        super().__init__(self.message)


class WorkbookNotFound(Exception):
    def __init__(self,
                 workbook_path: str,
                 classe: str,
                 metodo: str
                 ):
        self.message = (f"Objeto Workbook não encontrado.\n"
                        f"Verifique se o arquivo foi lido corretamente: '{workbook_path}'\n"
                        f"{'Classe: ' + classe if classe else ''}\n"
                        f"{'Método: ' + metodo if metodo else ''}"
                        )
        super().__init__(self.message)


class WorkSheetNotFound(Exception):
    def __init__(self,
                 workbook_path: str,
                 worksheet_name: str,
                 classe: str = '',
                 metodo: str = ''):
        self.message = (f"WorSheet não encontrada: '{worksheet_name}'.\n"
                        f"No arquivo: '{workbook_path}'.\n"
                        f"Verifique se o nome informado está correto.\n"
                        f"{'Classe: ' + classe if classe else ''}\n"
                        f"{'Método: ' + metodo if metodo else ''}"
                        )
        super().__init__(self.message)


class WorkSheetAlreadyExists(Exception):
    def __init__(self,
                 workbook_path: str,
                 worksheet_name: str,
                 classe: str = '',
                 metodo: str = ''
                 ):
        self.message = (f"WorSheet já existe: '{worksheet_name}'\n"
                        f"No arquivo: '{workbook_path}'.\n"
                        f"{'Classe: ' + classe if classe else ''}\n"
                        f"{'Método: ' + metodo if metodo else ''}"
                        )
        super().__init__(self.message)



if __name__ == '__main__':

    file_path_full = f'../Worksheet/Teste.xlsx'
    planilha = WorkBook(file_path=file_path_full)
    worksheet_name_test = 'Teste'
    planilha.readWorkSheet(worksheet_name_test)
    for row in planilha.worksheets[worksheet_name_test].iter_rows(values_only=True):
        print(row)

