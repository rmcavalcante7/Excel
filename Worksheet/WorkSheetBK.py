import traceback


from openpyxl.utils import column_index_from_string

from Projeto.Scripts.Auxiliar.String import TratarString
import inspect
from concurrent.futures import ThreadPoolExecutor, as_completed

from Projeto.Scripts.Excel.Excel import Excel
from Projeto.Scripts.Excel.Worksheet.WorksheetLayout import WorkSheetLayout
from Projeto.Scripts.Excel.Workbook.WorkBook import WorkBook


class WorkSheet:
    def __init__(self,
                 workbook: WorkBook,
                 worksheet_name: str,
                 worksheet_layout: [WorkSheetLayout | None]):
        self.workbook = workbook
        self.worksheet = self.workbook.readWorkSheet(worksheet_name)
        # self.base_name = base_name if base_name else 'worksheet_name'  # todo validar
        self.dict_base = {}
        self.worksheet_layout = worksheet_layout

    def __repr__(self):
        return (f'Classe: {self.__class__.__name__}\n'
                f'Worksheet Name: ({self.worksheet.title})')

    def validateWorkSheet(self):
        """Função utilizada para validar a estrutura da planilha.
        :return: None
        :raises: InvalidBaseLayout: Caso a estrutura da planilha não esteja de acordo com o padrão esperado.
        :raises: Exception: Caso ocorra algum erro inesperado.
        """
        nome_coluna_esperada = ''
        nome_col_base = ''
        pos_col = ''
        status_list = []
        column_row_number = 'Indefinido'
        try:
            column_row_number = self.worksheet_layout.data.get(self.worksheet_layout.columns_row_number_name)
            for row_number, row in enumerate(
                    self.worksheet.iter_rows(min_row=column_row_number,
                                             max_row=self.worksheet.max_row,
                                             values_only=True)):
                for key, values in self.worksheet_layout.data.items():
                    if key in self.worksheet_layout.lista_exececao:
                        continue

                    letra_coluna = values['col_origem'].upper() if values['col_origem'] else ''
                    if not letra_coluna:
                        continue
                    nome_coluna_esperada = ''
                    nome_col_base = ''
                    pos_col = f'{letra_coluna}{column_row_number}'
                    nome_coluna_esperada = TratarString.substituirCaractereEspecial(
                        TratarString.tratarEspacos(values['nome_inicial'])).upper()
                    if not nome_coluna_esperada:
                        continue
                    nome_col_base = TratarString.substituirCaractereEspecial(
                        TratarString.tratarEspacos(
                            str(row[column_index_from_string(letra_coluna) - 1]).upper()
                        )
                    ).upper()
                    if nome_coluna_esperada not in nome_col_base:
                        status_list.append(
                            f"A coluna '{nome_coluna_esperada}' não está na posição esperada: '{pos_col}'.\n")
                break

            if status_list:
                raise InvalidWorkSheetLayout(
                    file_path=self.workbook.file_path,
                    worksheet_name=self.worksheet.title,
                    worksheet_layout=self.worksheet_layout,
                    lista_erros=status_list,
                    classe=self.__class__.__name__,
                    metodo=inspect.currentframe().f_code.co_name)

                # raise InvalidBaseLayout(file_path=self.workbook.file_path,
                #                         worksheet_name=self.worksheet.title,
                #                         lista_erros=status_list,
                #                         classe=self.__class__.__name__,
                #                         metodo=inspect.currentframe().f_code.co_name)

        except InvalidWorkSheetLayout as ibl:
            raise ibl

        except Exception as e:
            # print(str(e))
            # print(traceback.format_exc())
            colunas_esperadas = [self.worksheet_layout.data["nome_inicial"]
                                 for key in self.worksheet_layout.data.keys()
                                 if key not in self.worksheet_layout.lista_exececao
                                 and self.worksheet_layout.data[key]["nome_inicial"]]
            raise Exception(
                f'Erro desconhecido no método {inspect.currentframe().f_code.co_name}() da classe {self.__class__.__name__}.\n'
                f'Nome da coluna esperada: "{nome_coluna_esperada}".\n'
                f'Nome da coluna na base: "{nome_col_base}".\n'
                f'Posição esperada da coluna na base: "{pos_col}".\n'
                f'Verifique se o arquivo está no padrão esperado.\n'
                f'Verifique se as colunas do arquivo estão na linha {column_row_number}.\n'
                f'Colunas esperadas: {colunas_esperadas}.\n'
                f'Arquivo em tratativa: "{self.workbook.file_path}"\n'
                f"Erros de layout encontrados:\n{'|'.join(status_list) if status_list else ''}\n"
                f'Exceção:\n{traceback.format_exc()}')

    def getDictBase(self, campo_chave=None):
        """Função utilizada para ler a base de dados e retornar um dicionário com os dados.
        :return: dict
        :raises: EmptyBase: Caso a base esteja vazia.
        :raises: Exception: Caso ocorra algum erro inesperado."""

        self.dict_base = {}

        # Função para processar uma linha da planilha
        def process_row(row, linha_id):
            row_result = {}

            work_sheet_layout = self.worksheet_layout.get_data()
            for field in work_sheet_layout:
                if field in self.worksheet_layout.lista_exececao:
                    continue
                coluna = work_sheet_layout.data[field]['col_origem']
                # row_result[field] = row[column_index_from_string(coluna) - 1]
                valor_celula = row[column_index_from_string(coluna) - 1]
                # row_result[field] = valor_celula if valor_celula is not None else ""
                work_sheet_layout.data[field]['valor'] = valor_celula if valor_celula else ""

            chave = work_sheet_layout.data[campo_chave]['valor'] if campo_chave else linha_id
            # row_result[chave] = self.worksheet_layout.data
            self.dict_base[chave] = work_sheet_layout.data

        try:
            with ThreadPoolExecutor() as executor:
                futures = []
                # Inicializa o contador de linha_id com o número da linha de dados
                # todo: estruturar para o linha_id pode ser a chave da linha da planilha
                linha_id = self.worksheet_layout.data.get(self.worksheet_layout.data_row_number_name)

                for row in self.worksheet.iter_rows(
                        min_row=self.worksheet_layout.data.get(self.worksheet_layout.data_row_number_name),
                        max_row=self.worksheet.max_row,
                        values_only=True):
                    futures.append(executor.submit(process_row, row, linha_id))
                    linha_id += 1  # Incrementa o contador

                for future in as_completed(futures):
                    future.result()

            if not self.dict_base:
                raise EmptyBase(file_path=self.workbook.file_path,
                                worksheet_name=self.worksheet.title,
                                classe=self.__class__.__name__,
                                metodo=inspect.currentframe().f_code.co_name
                                )

        except EmptyBase as eb:
            raise eb

        except Exception as e:
            print(traceback.format_exc())
            raise Exception(f"Erro não mapeado: {str(e)}")


    def getDictBase2(self, campo_chave=None):
        """Função utilizada para ler a base de dados e retornar um dicionário com os dados.
        :return: dict
        :raises: EmptyBase: Caso a base esteja vazia.
        :raises: Exception: Caso ocorra algum erro inesperado."""

        self.dict_base = {}

        # Função para processar uma linha da planilha
        def process_row(row, linha_id):
            row_result = {}

            for field in self.worksheet_layout.data:
                if field in self.worksheet_layout.lista_exececao:
                    continue
                coluna = self.worksheet_layout.data[field]['col_origem']
                # row_result[field] = row[column_index_from_string(coluna) - 1]
                valor_celula = row[column_index_from_string(coluna) - 1]
                row_result[field] = valor_celula if valor_celula is not None else ""

            chave = row_result.get(campo_chave) if campo_chave else linha_id
            if chave is not None:
                self.dict_base[chave] = row_result

        try:
            with ThreadPoolExecutor() as executor:
                futures = []
                # Inicializa o contador de linha_id com o número da linha de dados
                # todo: estruturar para o linha_id pode ser a chave da linha da planilha
                linha_id = self.worksheet_layout.data.get(self.worksheet_layout.data_row_number_name)

                for row in self.worksheet.iter_rows(
                        min_row=self.worksheet_layout.data.get(self.worksheet_layout.data_row_number_name),
                        max_row=self.worksheet.max_row,
                        values_only=True):
                    futures.append(executor.submit(process_row, row, linha_id))
                    linha_id += 1  # Incrementa o contador

                for future in as_completed(futures):
                    future.result()

            if not self.dict_base:
                raise EmptyBase(file_path=self.workbook.file_path,
                                worksheet_name=self.worksheet.title,
                                classe=self.__class__.__name__,
                                metodo=inspect.currentframe().f_code.co_name
                                )

        except EmptyBase as eb:
            raise eb

        except Exception as e:
            print(traceback.format_exc())
            raise Exception(f"Erro não mapeado: {str(e)}")

    def getDictBaseFromKey(self):
        """Função utilizada para ler a base de dados e retornar um dicionário com os dados.
        :return: dict
        :raises: EmptyBase: Caso a base esteja vazia.
        :raises: Exception: Caso ocorra algum erro inesperado."""

        self.dict_base = {}

        # Obtemos o índice da coluna que contém o 'cod_fornecedor'
        chave_principal = self.worksheet_layout.chave_principal
        coluna_chave = self.worksheet_layout.data[chave_principal]['col_origem']
        coluna_chave_index = column_index_from_string(coluna_chave) - 1  # Convertendo para índice de lista

        # Função para processar uma linha da planilha
        def process_row(row):
            row_result = {}

            for field in self.worksheet_layout.data:
                if field in self.worksheet_layout.lista_exececao:
                    continue
                coluna = self.worksheet_layout.data[field]['col_origem']
                row_result[field] = row[column_index_from_string(coluna) - 1]

            # Definir a chave principal com base no valor da coluna 'cod_fornecedor'
            chave = row[coluna_chave_index]

            if chave:  # Evitar adicionar chaves vazias
                self.dict_base[chave] = row_result  # Sobrescreve caso a chave já exista

        try:
            with ThreadPoolExecutor() as executor:
                futures = []
                for row in self.worksheet.iter_rows(
                        min_row=self.worksheet_layout.data_row_number,
                        max_row=self.worksheet.max_row,
                        values_only=True):
                    futures.append(executor.submit(process_row, row))

                for future in as_completed(futures):
                    future.result()

            if not self.dict_base:
                raise EmptyBase(file_path=self.workbook.file_path,
                                worksheet_name=self.worksheet.title,
                                classe=self.__class__.__name__,
                                metodo=inspect.currentframe().f_code.co_name
                                )

        except EmptyBase as eb:
            raise eb

        except Exception as e:
            print(traceback.format_exc())
            raise Exception(f"Erro não mapeado: {str(e)}")

    def getDictBaseOLD(self):
        """Função utilizada para ler a base de dados e retornar um dicionário com os dados.
        :return: dict
        :raises: EmptyBase: Caso a base esteja vazia.
        :raises: Exception: Caso ocorra algum erro inesperado."""

        # self.validarBase()
        self.dict_base = {}

        # Função para processar uma linha da planilha
        def process_row(row):
            chave_principal = row[0]
            chave_principal_tratada = TratarString.tratarChaveDicionario(chave_principal)
            row_result = {}

            if chave_principal_tratada not in self.dict_base:
                self.dict_base[chave_principal_tratada] = {}
                for field in self.worksheet_layout.data:
                    if field in self.worksheet_layout.lista_exececao:
                        continue
                    coluna = self.worksheet_layout.data[field]['col_origem']
                    self.dict_base[chave_principal_tratada][field] = row[column_index_from_string(coluna) - 1]

        try:
            # Usando ThreadPoolExecutor para processar as linhas em paralelo
            with ThreadPoolExecutor() as executor:
                futures = []
                for row in self.worksheet.iter_rows(
                        min_row=self.worksheet_layout.data.get(self.worksheet_layout.data_row_number_name),
                        max_row=self.worksheet.max_row,
                        values_only=True):
                    # Submete o processamento de cada linha para a execução em paralelo
                    futures.append(executor.submit(process_row, row))

                # Aguardando a finalização de todas as threads
                for future in as_completed(futures):
                    future.result()  # Isso garante que qualquer exceção será levantada aqui, se houver

            if not self.dict_base:
                raise EmptyBase(file_path=self.workbook.file_path,
                                worksheet_name=self.worksheet.title,
                                classe=self.__class__.__name__,
                                metodo=inspect.currentframe().f_code.co_name
                                )

        except EmptyBase as eb:
            raise eb

        except Exception as e:
            print(traceback.format_exc())
            raise Exception(f"Erro não mapeado: {str(e)}")



class InvalidWorkSheetLayout(Exception):
    def __init__(self,
                 file_path: str,
                 worksheet_name: str,
                 worksheet_layout: WorkSheetLayout,
                 lista_erros: list,
                 classe: str = '',
                 metodo: str = ''
                 ):
        colunas_esperadas = '\n'.join([(f"Coluna '{worksheet_layout.data[col]['nome_inicial']}' na posição "
                              f"'{worksheet_layout.data[col]['col_origem']}{worksheet_layout.data[worksheet_layout.columns_row_number_name]}'.")
                             for col in worksheet_layout.data.keys()
                             if col not in worksheet_layout.lista_exececao and worksheet_layout.data[col]["nome_inicial"]])

        self.message = f"A aba '{worksheet_name}' da base fixa '{file_path}' não está no padrão esperado.\n" \
                       f"Padrão Esperado:\n{colunas_esperadas}\n\n" \
                       f"Erros encontrados:\n{''.join(lista_erros)}\n" \
                       f"Classe '{classe}'.\n" \
                       f"Método '{metodo}()'"
        super().__init__(self.message)

class EmptyBase(Exception):
    def __init__(self,
                 file_path: str,
                 worksheet_name: str,
                 classe: str = '',
                 metodo: str = ''
                 ):
        self.message = (
            f"Base '{worksheet_name}' está vazia.\n"
            f"Verifique se não houve alterações no padrão da aba '{worksheet_name}' da base fixa "
            f"'{file_path}'."
            f"Classe '{classe}'.\n"
            f"Método '{metodo}()'"
        )
        super().__init__(self.message)


if '__main__' == __name__:
    from Projeto.Scripts.Excel.Worksheet.LayoutTeste import LayoutTESTE
    from pprint import pprint

    layout_teste = LayoutTESTE()
    file_path_teste = f'Teste.xlsx'
    planilha = WorkBook(excel=Excel(),
                        file_path=file_path_teste,
                        read_only=True,
                        data_only=True)
    worksheet_bloqueio_name = 'Teste'

    # Criação do objeto WorkSheet
    work_sheet = WorkSheet(
        workbook=planilha,
        worksheet_name=worksheet_bloqueio_name,
        worksheet_layout=layout_teste
    )

    work_sheet.validateWorkSheet()
    work_sheet.getDictBase()
    pprint(work_sheet.dict_base)
    #